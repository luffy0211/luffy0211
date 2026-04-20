import asyncio
import os
import random
import requests
import json
import time
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
import openpyxl
from datetime import datetime
from pathlib import Path

def read_url_from_excel(filename='urls.xlsx', row=1, col=1):
    """从Excel文件中读取URL"""
    from pathlib import Path
    
    desktop_path = str(Path.home() / "Desktop")
    filepath = os.path.join(desktop_path, filename)
    
    try:
        if not os.path.exists(filepath):
            print(f"错误：文件 {filepath} 不存在！")
            return None
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        cell_value = ws.cell(row=row, column=col).value
        
        if cell_value:
            print(f"从Excel第{row}行第{col}列读取到URL: {cell_value}")
            return cell_value
        else:
            print(f"错误：第{row}行第{col}列为空！")
            return None
            
    except Exception as e:
        print(f"读取Excel时出错: {e}")
        return None

def write_to_excel(data, product_data=None, image_paths=None, filename='taobao_products.xlsx'):
    """将数据写入Excel文件（覆盖写入，只保留最新数据）"""
    from pathlib import Path
    
    desktop_path = str(Path.home() / "Desktop")
    filename = os.path.join(desktop_path, filename)
    
    # 如果文件存在，删除旧文件
    if os.path.exists(filename):
        try:
            os.remove(filename)
        except:
            pass
    
    # 创建新文件
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['商品链接', '商品名称', '价格', '店铺名称', '', '风格', '颜色分类', '适用季节', '材质成分', '面料', '安全等级', '', '', '主图', '', '', '', 'SKU图片', '身高']
    ws.append(headers)
    
    # 从product_data中提取属性
    attrs = product_data.get('attributes', []) if product_data else []
    
    style = ''
    fabric = ''
    season = ''
    material = ''
    safety = ''
    height = ''
    color_str = ''
    
    # 从attributes中匹配所有字段
    # 淘宝属性格式为 "key\nvalue"，需要分割取值
    for attr in attrs:
        # 按换行符分割，取值部分
        parts = attr.split('\n', 1)
        if len(parts) == 2:
            key = parts[0].strip()
            value = parts[1].strip()
        else:
            # 兼容 "key： value" 格式（3e3e）
            if '：' in attr:
                key, value = attr.split('：', 1)
                key = key.strip()
                value = value.strip()
            elif ':' in attr:
                key, value = attr.split(':', 1)
                key = key.strip()
                value = value.strip()
            else:
                continue
        
        # 风格
        if key == '风格':
            style = value
        # 面料
        elif key == '面料':
            fabric = value
        # 适用季节
        elif key == '适用季节':
            season = value
        # 材质成分
        elif key == '材质成分':
            material = value
        # 安全等级/安全类别
        elif key in ('安全等级', '安全类别'):
            safety = value
        # 身高
        elif key == '身高':
            height = value
        # 颜色分类
        elif key == '颜色分类':
            color_str = value
    
    # 主图和SKU图路径只保留文件夹路径（从第一个文件路径提取）
    main_path = image_paths.get('main', '') if image_paths else ''
    sku_path = image_paths.get('sku', '') if image_paths else ''
    
    # 从第一个文件路径提取文件夹
    main_folder = os.path.dirname(main_path.split(';')[0].strip()) if main_path else ''
    sku_folder = os.path.dirname(sku_path.split(';')[0].strip()) if sku_path else ''
    
    # 商品名称去掉"复制"和空格
    title = data.get('商品名称', '').replace('复制', '').replace(' ', '').strip()
    
    row_data = [
        data.get('商品链接', ''),
        title,
        data.get('价格', ''),
        data.get('店铺名称', ''),
        '',
        style,
        color_str,
        season,
        material,
        fabric,
        safety,
        '',
        '',
        main_folder,
        '',
        '',
        '',
        sku_folder,
        height,
    ]
    ws.append(row_data)
    wb.save(filename)
    print(f"数据已保存到: {filename}")

async def random_delay(min_sec=1, max_sec=3):
    """随机延迟，模拟人类行为"""
    delay = random.uniform(min_sec, max_sec)
    await asyncio.sleep(delay)

async def download_images(product_name, main_images, sku_images):
    """下载图片到桌面童装文件夹"""
    from pathlib import Path
    
    image_paths = {'main': '', 'sku': ''}
    desktop_path = r"C:\Users\郭雷A\Desktop"
    
    safe_name = "".join(c for c in product_name if c not in r'\/:*?"<>|').strip()[:50]
    if not safe_name:
        safe_name = "product"
    
    product_folder = os.path.join(desktop_path, '童装', safe_name)
    main_folder = product_folder
    sku_folder = os.path.join(product_folder, 'sku')
    
    os.makedirs(main_folder, exist_ok=True)
    os.makedirs(sku_folder, exist_ok=True)
    
    print(f"图片保存目录: {product_folder}")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://detail.tmall.com/',
    }
    
    main_paths = []
    sku_paths = []
    
    # 下载主图
    print(f"正在下载主图...")
    for i, src in enumerate(main_images):
        try:
            print(f"  正在下载主图 {i+1}/{len(main_images)}...")
            response = requests.get(src, timeout=15, headers=headers)
            
            if response.status_code == 200:
                ext = '.jpg'
                if '.png' in src:
                    ext = '.png'
                
                filename = f"主图_{i+1}{ext}"
                filepath = os.path.join(main_folder, filename)
                main_paths.append(filepath)
                
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                
                print(f"    已保存: {filename}")
            else:
                print(f"    下载失败，状态码: {response.status_code}")
            
            await asyncio.sleep(random.uniform(0.3, 0.8))
            
        except Exception as e:
            print(f"    下载失败: {e}")
            continue
    
    # 下载SKU图
    print(f"正在下载SKU图...")
    for i, src in enumerate(sku_images):
        try:
            print(f"  正在下载SKU图 {i+1}/{len(sku_images)}...")
            response = requests.get(src, timeout=15, headers=headers)
            
            if response.status_code == 200:
                ext = '.jpg'
                if '.png' in src:
                    ext = '.png'
                
                filename = f"SKU_{i+1}{ext}"
                filepath = os.path.join(sku_folder, filename)
                sku_paths.append(filepath)
                
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                
                print(f"    已保存: {filename}")
            else:
                print(f"    下载失败，状态码: {response.status_code}")
            
            await asyncio.sleep(random.uniform(0.3, 0.8))
            
        except Exception as e:
            print(f"    下载失败: {e}")
            continue
    
    image_paths['main'] = '; '.join(main_paths)
    image_paths['sku'] = '; '.join(sku_paths)
    
    print(f"图片下载完成！主图: {len(main_paths)}张, SKU图: {len(sku_paths)}张")
    
    return image_paths

async def crawl_taobao_product(page, product_url):
    """参照3e3e.py结构抓取淘宝商品信息"""
    data = {}
    
    try:
        # 1. 商品标题
        try:
            title_el = await page.query_selector('span.mainTitle--R75fTcZL')
            data['title'] = (await title_el.inner_text()).strip() if title_el else ""
        except:
            data['title'] = ""
        
        # 2. 价格
        try:
            price_el = await page.query_selector('div.highlightPrice--LlVWiXXs span.text--LP7Wf49z')
            if price_el:
                data['price'] = (await price_el.inner_text()).strip()
            else:
                price_el2 = await page.query_selector('span.text--jyiUrkMu')
                data['price'] = (await price_el2.inner_text()).strip() if price_el2 else ""
        except:
            data['price'] = ""
        
        # 3. 主图和SKU图片
        main_imgs = []
        sku_imgs = []
        
        try:
            main_images = await page.query_selector_all('xpath=//img[contains(@class, "thumbnailPic--QasTmWDm")]')
            for img in main_images:
                src = await img.get_attribute('src')
                if src:
                    if src.startswith('//'):
                        src = 'https:' + src
                    # 处理主图后缀 .jpg
                    if '.jpg' in src and '.webp' in src:
                        src = src.replace('.jpg_.webp', '.jpg').replace('.jpg_.webp', '')
                    if '_q50' in src:
                        src = src.split('_q50')[0]
                    if src and src not in main_imgs:
                        main_imgs.append(src)
        except:
            pass
        
        try:
            sku_images = await page.query_selector_all('xpath=//img[contains(@class, "valueItemImg--GC9bH5my")]')
            for img in sku_images:
                src = await img.get_attribute('src')
                if src:
                    if src.startswith('//'):
                        src = 'https:' + src
                    # 处理SKU图后缀 jpg_.webp -> jpg
                    if 'jpg_.webp' in src:
                        src = src.replace('jpg_.webp', 'jpg')
                    if '_30x30' in src:
                        src = src.replace('_30x30', '_200x200')
                    if src and src not in sku_imgs:
                        sku_imgs.append(src)
        except:
            pass
        
        data['main_images'] = main_imgs
        data['sku_images'] = sku_imgs
        data['images'] = main_imgs + sku_imgs
        
        # 4. 商品视频
        data['video_url'] = ""
        
        # 5. 颜色分类
        colors = []
        try:
            color_items = await page.query_selector_all('xpath=//li[contains(@class, "sku-item")]')
            for item in color_items:
                color = await item.get_attribute('data-color')
                if color and color not in colors:
                    colors.append(color)
        except:
            pass
        data['colors'] = colors
        
        # 6. 尺码
        sizes = []
        try:
            size_items = await page.query_selector_all('xpath=//li[contains(@class, "sku-size-item")]')
            for item in size_items:
                size = await item.inner_text()
                if size and size.strip() not in sizes:
                    sizes.append(size.strip())
        except:
            pass
        data['sizes'] = sizes
        
        # 7. 店铺信息
        try:
            shop_el = await page.query_selector('a.shopName---gV6rjH4')
            data['shop_name'] = (await shop_el.inner_text()).strip() if shop_el else ""
        except:
            data['shop_name'] = ""
        
        try:
            shop_el = await page.query_selector('a.shopName---gV6rjH4')
            data['shop_url'] = await shop_el.get_attribute('href') if shop_el else ""
        except:
            data['shop_url'] = ""
        
        try:
            address_el = await page.query_selector('div.shop-address')
            data['shop_address'] = (await address_el.inner_text()).strip() if address_el else ""
        except:
            data['shop_address'] = ""
        
        # 8. 商品属性
        attrs = []
        try:
            param_els = await page.query_selector_all('div.generalParamsInfoItem--qLqLDVWp')
            for el in param_els:
                text = await el.inner_text()
                if text:
                    attrs.append(text.strip())
        except:
            pass
        
        try:
            emphasis_els = await page.query_selector_all('div.emphasisParamsInfoItem--H5Qt3iog')
            for el in emphasis_els:
                text = await el.inner_text()
                if text:
                    attrs.append(text.strip())
        except:
            pass
        
        data['attributes'] = attrs
        
        # 9. 商品ID
        try:
            collect_btn = await page.query_selector('.product-collect-btn')
            data['product_id'] = await collect_btn.get_attribute('monitor-productid') if collect_btn else ""
        except:
            data['product_id'] = ""
        
        # 10. 商品链接
        data['product_url'] = product_url
        
    except Exception as e:
        print(f"抓取商品信息时出错: {e}")
    
    return data

async def run_crawler():
    state_file = "D:\\zsh\\taobao_state.json"
    
    if not os.path.exists(state_file):
        print(f"错误：未找到 {state_file}，请先运行保存登录状态的脚本。")
        return

    stealth = Stealth(
        navigator_webdriver=False,
        chrome_app=False,
        chrome_csi=False,
        chrome_load_times=False,
    )
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False, 
            channel="msedge",
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-infobars',
                '--window-size=1920,1080',
            ]
        )
        
        context = await browser.new_context(
            storage_state=state_file,
            viewport={'width': 1920, 'height': 1080},
            locale='zh-CN',
            timezone_id='Asia/Shanghai',
            permissions=['geolocation'],
            extra_http_headers={
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            }
        )
        page = await context.new_page()
        
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            Object.defineProperty(navigator, 'languages', {
                get: () => ['zh-CN', 'zh', 'en']
            });
            const originalQuery = window.navigator.permissions.query;
            window.navigator.permissions.query = (parameters) => (
                parameters.name === 'notifications' ?
                    Promise.resolve({ state: Notification.permission }) :
                    originalQuery(parameters)
            );
            if (window.chrome) {
                window.chrome.runtime = {
                    connect: function() {},
                    sendMessage: function() {}
                };
            }
        """)
        
        await stealth.apply_stealth_async(page)
        await random_delay(1, 3)

        product_url = read_url_from_excel(filename='urls.xlsx', row=1, col=1)
        
        if not product_url:
            print("错误：无法获取商品URL，程序退出。")
            await browser.close()
            return
        
        print(f"正在访问天猫商品详情页...")
        await random_delay(1, 2)
        
        try:
            await page.goto(product_url, wait_until="domcontentloaded", timeout=30000)
            await random_delay(2, 4)
            await page.mouse.move(random.randint(300, 800), random.randint(200, 600))
            await random_delay(0.5, 1.5)
            
            title_selector = 'span.mainTitle--R75fTcZL'
            
            print("正在检测页面内容...")
            try:
                await page.wait_for_selector(title_selector, timeout=15000)
            except:
                print("未发现商品标题，可能出现滑块验证，请在浏览器中手动操作...")
                while True:
                    if await page.query_selector(title_selector):
                        break
                    await asyncio.sleep(1)

            print("\n" + "="*50)
            
            # 调用抓取函数
            data = await crawl_taobao_product(page, product_url)
            
            # 打印抓取结果
            print(f"【商品名称】: {data.get('title', '')}")
            print(f"【价格】: {data.get('price', '')}")
            print(f"【店铺名称】: {data.get('shop_name', '')}")
            print(f"【图片数量】: {len(data.get('images', []))}")
            print(f"【颜色数量】: {len(data.get('colors', []))}")
            print(f"【尺码数量】: {len(data.get('sizes', []))}")
            print(f"【属性数量】: {len(data.get('attributes', []))}")
            
            if data.get('attributes'):
                print("\n【商品属性】:")
                for attr in data.get('attributes', []):
                    print(f"  {attr}")
            
            # 抓取图片
            print("\n正在处理商品图片...")
            image_paths = await download_images(
                data.get('title', 'product'),
                data.get('main_images', []),
                data.get('sku_images', [])
            )
            
            # 商品名称去掉"复制"和空格
            title = data.get('title', '').replace('复制', '').replace(' ', '').strip()
            
            # 写入Excel
            write_to_excel({
                '商品名称': title,
                '价格': data.get('price', ''),
                '商品链接': data.get('product_url', ''),
                '店铺名称': data.get('shop_name', ''),
            }, data, image_paths)
            
            # 保存JSON文件
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            product_id = data.get('product_id', 'unknown')
            filename = f"taobao_product_{product_id}_{timestamp}.json"
            
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            print(f"\nJSON数据已保存到: {filename}")
            print("="*50)

        except Exception as e:
            print(f"抓取时发生错误: {e}")
        
        finally:
            print("脚本执行完毕，5秒后关闭...")
            await asyncio.sleep(5)
            await browser.close()

if __name__ == "__main__":
    asyncio.run(run_crawler())