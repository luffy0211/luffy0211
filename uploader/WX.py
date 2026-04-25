import asyncio
import os
import json
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl
import time

def read_excel_data(filename='影刀上架参数.xlsx'):
    """从桌面Excel读取上传参数"""
    desktop_path = str(Path.home() / "Desktop")
    filepath = os.path.join(desktop_path, filename)
    
    if not os.path.exists(filepath):
        print(f"错误：找不到文件 {filepath}")
        return None
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    data = []
    for row in range(2, ws.max_row + 1):
        style = ws.cell(row=row, column=6).value    # F列：风格
        color = ws.cell(row=row, column=7).value    # G列：颜色
        material = ws.cell(row=row, column=9).value  # I列：面料材质
        image_path = ws.cell(row=row, column=14).value  # N列
        title = ws.cell(row=row, column=15).value       # O列
        if image_path or title:
            data.append({'style': style, 'color': color, 'material': material, 'image_path': image_path, 'title': title})
    
    print(f"从Excel读取了 {len(data)} 条数据")
    return data

def get_images_from_folder(folder_path):
    """获取文件夹下的所有图片文件"""
    if not folder_path or not os.path.exists(folder_path):
        return []
    
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp']
    images = []
    
    if os.path.isdir(folder_path):
        for file in os.listdir(folder_path):
            if os.path.splitext(file)[1].lower() in image_extensions:
                images.append(os.path.abspath(os.path.join(folder_path, file)))
    elif os.path.isfile(folder_path):
        images.append(os.path.abspath(folder_path))
    
    return sorted(images)

async def handle_popups(page):
    """深度清理弹窗（检查主文档和Shadow DOM）"""
    print("正在扫描并清理弹窗...")
    
    popup_js = """
    () => {
        const findAndClick = (root) => {
            const btns = root.querySelectorAll('button, .weui-desktop-btn, .weui-btn');
            for (const btn of btns) {
                const text = btn.innerText || '';
                if (text.includes('我知道了') || text.includes('确定') || text.includes('知道了')) {
                    btn.click();
                    return true;
                }
            }
            return false;
        };
        
        // 1. 检查主文档
        let found = findAndClick(document);
        
        // 2. 检查 micro-app Shadow DOM
        const microApp = document.querySelector('micro-app[name="goods"]');
        if (microApp && microApp.shadowRoot) {
            if(findAndClick(microApp.shadowRoot)) found = true;
        }
        
        return found;
    }
    """
    
    for _ in range(3): # 尝试3次
        if await page.evaluate(popup_js):
            print("成功点击了一个弹窗按钮")
            await asyncio.sleep(1)
        else:
            break

async def test_weixin_login():
    state_file = "weixin_state.json"
    
    if not os.path.exists(state_file):
        print(f"错误：找不到 {state_file}")
        return
    
    excel_data = read_excel_data()
    if not excel_data: return

    async with async_playwright() as p:
        # 1. 启动时通过参数实现窗口最大化
        browser = await p.chromium.launch(
            headless=False, 
            channel="msedge",
            args=["--start-maximized"] 
        )
        
        # 2. 创建上下文时禁用默认视口，允许窗口最大化生效
        context = await browser.new_context(
            no_viewport=True, 
            storage_state=state_file
        )
        
        page = await context.new_page()

        # 进入商品发布页
        await page.goto("https://store.weixin.qq.com/shop/goods/entry")
        print("等待页面及微应用加载...")
        
        # 初始等待与处理弹窗
        await asyncio.sleep(5)
        await handle_popups(page)
        
        # 等待微应用加载完成
        try:
            await page.wait_for_selector('micro-app[name="goods"]', timeout=20000)
            print("微应用 goods 已就绪")
        except:
            print("微应用加载超时，请检查登录状态")
            return

        # 只处理第一条数据
        if excel_data:
            item = excel_data[0]
            idx = 0
            print(f"\n--- 正在处理第 {idx + 1} 条 (仅处理第一条) ---")
            image_path = item.get('image_path')
            title = item.get('title')
            material = item.get('material')  # 面料材质
            color = item.get('color')  # 颜色
            style = item.get('style')  # 风格

            print(f"图片路径: {image_path}")
            print(f"标题: {title}")
            print(f"面料材质: {material}")
            print(f"颜色: {color}")
            print(f"风格: {style}")

            # 1. 准备图片路径
            image_files = get_images_from_folder(str(image_path))
            if not image_files:
                print(f"警告：路径 {image_path} 下未找到图片")
            else:
                # 2. 上传图片 (核心修复部分)
                try:
                    print(f"准备上传 {len(image_files)} 张图片...")
                    
                    # 等待微应用完全加载
                    await page.wait_for_timeout(3000)
                    
                    # 关键：先设置文件选择器监听器，再点击
                    async with page.expect_file_chooser(timeout=15000) as fc_info:
                        print("正在查找并点击上传按钮...")
                        
                        # 使用统一的选择器策略 - 查找上传按钮
                        await page.evaluate("""
                            () => {
                                const microApp = document.querySelector('micro-app[name="goods"]');
                                if (!microApp || !microApp.shadowRoot) {
                                    console.log('No micro-app or shadow root');
                                    return false;
                                }
                                
                                const shadowRoot = microApp.shadowRoot;
                                
                                // 统一的选择器列表
                                const selectors = [
                                    '.picture_add_content',
                                    '.icon_add',
                                    '.goods-image-upload',
                                    '[class*="upload"]',
                                    '[class*="picture"]',
                                    '[class*="add"]'
                                ];
                                
                                for (const sel of selectors) {
                                    const elements = shadowRoot.querySelectorAll(sel);
                                    for (const element of elements) {
                                        if (element.offsetWidth > 0 && element.offsetHeight > 0) {
                                            console.log('Found upload area:', sel);
                                            element.click();
                                            return true;
                                        }
                                    }
                                }
                                
                                console.log('No upload area found');
                                return false;
                            }
                        """)
                    
                    # 获取文件选择器并设置文件
                    file_chooser = await fc_info.value
                    await file_chooser.set_files(image_files)
                    print("图片上传指令已发送")
                    
                    # 等待图片上传完成
                    await page.wait_for_timeout(5000)
                    
                except Exception as e:
                    print(f"图片上传失败: {e}")
                    print("请手动点击上传...")
                    input("上传完成后按回车继续...")

                # 3. 填写标题
                if title:
                    try:
                        print("正在查找标题输入框...")
                        
                        # 使用统一的选择器策略 - 查找标题输入框
                        fill_result = await page.evaluate("""
                            (titleText) => {
                                const microApp = document.querySelector('micro-app[name="goods"]');
                                if (!microApp || !microApp.shadowRoot) return { success: false, reason: 'no shadow root' };
                                
                                const shadowRoot = microApp.shadowRoot;
                                
                                // 统一的选择器列表
                                const selectors = [
                                    'input[placeholder*="商品名称"]',
                                    'input[placeholder*="关键字"]',
                                    'input[placeholder*="商品标题"]',
                                    'input[placeholder*="标题"]',
                                    'input[data-eleid="product_title"]',
                                    '.weui-desktop-form__input'
                                ];
                                
                                for (const sel of selectors) {
                                    const input = shadowRoot.querySelector(sel);
                                    if (input) {
                                        console.log('Found title input:', sel);
                                        input.value = titleText;
                                        input.dispatchEvent(new Event('input', { bubbles: true }));
                                        input.dispatchEvent(new Event('change', { bubbles: true }));
                                        return { success: true, selector: sel };
                                    }
                                }
                                
                                return { success: false, reason: 'no title input found' };
                            }
                        """, title)
                        
                        if fill_result.get('success'):
                            print(f"标题已填写: {title} (选择器: {fill_result.get('selector')})")
                        else:
                            print(f"填写标题失败: {fill_result.get('reason')}")
                            print("请手动填写标题...")
                            input("填写完成后按回车继续...")
                            
                    except Exception as e:
                        print(f"标题填写失败: {e}")
                        print("请手动填写标题...")
                        input("填写完成后按回车继续...")
                    
                    # 4. 点击下一步
                    try:
                        print("正在查找并点击下一步按钮...")
                        
                        await page.wait_for_timeout(2000)
                        
                        # 检查是否出现推荐理由
                        print("检查是否出现推荐理由...")
                        
                        # 等待推荐理由出现
                        recommend_reason_found = False
                        for i in range(10):  # 最多等待10次，每次2秒
                            found = await page.evaluate("""
                                () => {
                                    const microApp = document.querySelector('micro-app[name="goods"]');
                                    if (!microApp || !microApp.shadowRoot) return false;
                                    
                                    const shadowRoot = microApp.shadowRoot;
                                    
                                    // 查找推荐理由元素
                                    const recommendElements = shadowRoot.querySelectorAll('.recommend-reason, [class*="recommend"]');
                                    for (const element of recommendElements) {
                                        const text = element.textContent || element.innerText || '';
                                        if (text.includes('推荐理由')) {
                                            console.log('Found recommend reason:', text);
                                            return true;
                                        }
                                    }
                                    
                                    // 兜底：查找包含推荐理由文本的元素
                                    const elements = shadowRoot.querySelectorAll('div');
                                    for (const element of elements) {
                                        const text = element.textContent || element.innerText || '';
                                        if (text.includes('推荐理由')) {
                                            console.log('Found recommend reason by text:', text);
                                            return true;
                                        }
                                    }
                                    
                                    return false;
                                }
                            """)
                            
                            if found:
                                recommend_reason_found = True
                                print("推荐理由已出现")
                                break
                            
                            print(f"等待推荐理由出现... (第{i+1}/10)")
                            await page.wait_for_timeout(2000)
                        
                        if not recommend_reason_found:
                            print("未找到推荐理由，继续执行")
                        
                        # 点击下一步
                        click_result = await page.evaluate("""
                            () => {
                                const microApp = document.querySelector('micro-app[name="goods"]');
                                if (!microApp || !microApp.shadowRoot) return false;
                                
                                const shadowRoot = microApp.shadowRoot;
                                
                                // 统一的选择器列表
                                const selectors = [
                                    'button.weui-desktop-btn_primary',
                                    '.button_content button',
                                    '.weui-desktop-btn_wrp button',
                                    '.fix_box button',
                                    'button:contains("下一步")'
                                ];
                                
                                // 等待按钮变为可用
                                let button = null;
                                let attempts = 0;
                                const maxAttempts = 30; // 最多尝试30次
                                
                                while (attempts < maxAttempts) {
                                    for (const sel of selectors) {
                                        const buttons = shadowRoot.querySelectorAll(sel);
                                        for (const btn of buttons) {
                                            const text = btn.textContent || btn.innerText || '';
                                            if (text.includes('下一步')) {
                                                button = btn;
                                                break;
                                            }
                                        }
                                        if (button) break;
                                    }
                                    
                                    // 检查按钮是否可用
                                    if (button && !button.className.includes('weui-desktop-btn_disabled')) {
                                        console.log('Next button is enabled, clicking...');
                                        button.click();
                                        return true;
                                    }
                                    
                                    attempts++;
                                    // 等待100ms后重试
                                    for (let i = 0; i < 100; i++) {
                                        if (button && !button.className.includes('weui-desktop-btn_disabled')) {
                                            break;
                                        }
                                        const start = Date.now();
                                        while (Date.now() - start < 1) {
                                            // 空循环
                                        }
                                    }
                                }
                                
                                console.log('Next button not enabled after', maxAttempts, 'attempts');
                                return false;
                            }
                        """)
                        
                        if click_result:
                            print("已点击下一步按钮")
                        else:
                            print("未找到下一步按钮")
                            print("请手动点击下一步...")
                            input("点击完成后按回车继续...")
                            
                    except Exception as e:
                        print(f"点击下一步失败: {e}")
                        print("请手动点击下一步...")
                        input("点击完成后按回车继续...")

                    # 5. 填写面料材质
                    if material:
                        try:
                            print("正在查找并填写面料材质...")
                            
                            # 等待页面加载
                            await page.wait_for_timeout(3000)
                            
                            # 使用统一的选择器策略 - 查找面料材质输入框
                            fill_result = await page.evaluate("""
                                (materialText) => {
                                    const microApp = document.querySelector('micro-app[name="goods"]');
                                    if (!microApp || !microApp.shadowRoot) return { success: false, reason: 'no shadow root' };
                                    
                                    const shadowRoot = microApp.shadowRoot;
                                    
                                    // 统一的选择器列表
                                    const selectors = [
                                        'input[placeholder*="面料材质"]',
                                        '.main-option input',
                                        '.weui-desktop-form__input[placeholder*="面料"]'
                                    ];
                                    
                                    for (const sel of selectors) {
                                        const input = shadowRoot.querySelector(sel);
                                        if (input) {
                                            console.log('Found material input:', sel);
                                            input.value = materialText;
                                            input.dispatchEvent(new Event('input', { bubbles: true }));
                                            input.dispatchEvent(new Event('change', { bubbles: true }));
                                            return { success: true, selector: sel };
                                        }
                                    }
                                    
                                    return { success: false, reason: 'no material input found' };
                                }
                            """, material)
                            
                            if fill_result.get('success'):
                                print(f"面料材质已填写: {material} (选择器: {fill_result.get('selector')})")
                            else:
                                print(f"填写面料材质失败: {fill_result.get('reason')}")
                                print("请手动填写面料材质...")
                                input("填写完成后按回车继续...")
                                
                        except Exception as e:
                            print(f"填写面料材质失败: {e}")
                            print("请手动填写面料材质...")
                            input("填写完成后按回车继续...")

                    # 6. 填写颜色
                    if color:
                        try:
                            print("正在查找并填写颜色...")
                            
                            # 使用统一的选择器策略 - 查找颜色输入框
                            fill_result = await page.evaluate("""
                                (colorText) => {
                                    const microApp = document.querySelector('micro-app[name="goods"]');
                                    if (!microApp || !microApp.shadowRoot) return { success: false, reason: 'no shadow root' };
                                    
                                    const shadowRoot = microApp.shadowRoot;
                                    
                                    // 统一的选择器列表
                                    const selectors = [
                                        'input[placeholder*="颜色"]',
                                        '.main-option input[placeholder*="颜色"]'
                                    ];
                                    
                                    for (const sel of selectors) {
                                        const input = shadowRoot.querySelector(sel);
                                        if (input) {
                                            console.log('Found color input:', sel);
                                            input.value = colorText;
                                            input.dispatchEvent(new Event('input', { bubbles: true }));
                                            input.dispatchEvent(new Event('change', { bubbles: true }));
                                            return { success: true, selector: sel };
                                        }
                                    }
                                    
                                    return { success: false, reason: 'no color input found' };
                                }
                            """, color)
                            
                            if fill_result.get('success'):
                                print(f"颜色已填写: {color} (选择器: {fill_result.get('selector')})")
                            else:
                                print(f"填写颜色失败: {fill_result.get('reason')}")
                                print("请手动填写颜色...")
                                
                                
                        except Exception as e:
                            print(f"填写颜色失败: {e}")
                            print("请手动填写颜色...")
                            

                    # 7. 填写风格
                    if style:
                        try:
                            print("正在查找并选择风格...")
                            
                            # 使用统一的选择器策略 - 处理风格下拉框
                            select_result = await page.evaluate("""
                                (styleText) => {
                                    const microApp = document.querySelector('micro-app[name="goods"]');
                                    if (!microApp || !microApp.shadowRoot) return { success: false, reason: 'no shadow root' };
                                    
                                    const shadowRoot = microApp.shadowRoot;
                                    
                                    // 1. 点击风格下拉框
                                    const dropdown = shadowRoot.querySelector('input[placeholder*="风格"]');
                                    if (dropdown) {
                                        dropdown.click();
                                        console.log('Clicked style dropdown');
                                    }
                                    
                                    // 2. 等待下拉菜单出现
                                    setTimeout(() => {
                                        // 3. 查找并点击匹配的风格选项
                                        const options = shadowRoot.querySelectorAll('.weui-desktop-dropdown__list-ele');
                                        let found = false;
                                        
                                        for (const option of options) {
                                            const text = option.textContent || option.innerText || '';
                                            if (text.includes(styleText)) {
                                                option.click();
                                                console.log('Selected style:', styleText);
                                                found = true;
                                                break;
                                            }
                                        }
                                        
                                        // 如果没找到，选择第一个选项
                                        if (!found && options.length > 0) {
                                            options[0].click();
                                            console.log('Selected first style option');
                                        }
                                    }, 500);
                                    
                                    return { success: true };
                                }
                            """, style)
                            
                            if select_result.get('success'):
                                print(f"风格已选择: {style}")
                            else:
                                print(f"选择风格失败: {select_result.get('reason')}")
                                print("请手动选择风格...")
                                
                                
                        except Exception as e:
                            print(f"选择风格失败: {e}")
                            print("请手动选择风格...")
                            

                # 处理完留一点缓冲时间
                await page.wait_for_timeout(3000)

                

        print("\n所有任务处理完毕！")
        # 等待用户确认
        await browser.close()

if __name__ == "__main__":
    asyncio.run(test_weixin_login())