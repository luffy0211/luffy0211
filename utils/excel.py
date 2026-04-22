import os
from typing import Optional
import openpyxl
from utils.logger import setup_logger

logger = setup_logger("excel")


def get_desktop_path() -> str:
    from pathlib import Path
    return str(Path.home() / "Desktop")


def read_urls(filepath: str) -> list[str]:
    """从 Excel 读取所有 URL（A列，跳过空行）"""
    if not os.path.exists(filepath):
        logger.error(f"文件不存在: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val and str(val).strip().startswith("http"):
            urls.append(str(val).strip())
    wb.close()
    logger.info(f"从 {filepath} 读取了 {len(urls)} 条 URL")
    return urls


def read_upload_data(filepath: str, columns: dict) -> list[dict]:
    """从上架 Excel 读取商品数据

    columns: {'style': 6, 'color': 7, 'material': 9, 'image_path': 14, 'title': 15, ...}
    
    支持的字段：
    - colors: 逗号分隔的颜色列表，如 "白色,黑色,红色"
    - sizes: 逗号分隔的尺码列表，如 "S,M,L,XL"
    - sku_images: JSON格式或文件夹路径
    """
    if not os.path.exists(filepath):
        logger.error(f"文件不存在: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    data = []

    for row in range(2, ws.max_row + 1):
        item = {}
        for key, col in columns.items():
            cell_value = ws.cell(row=row, column=col).value
            
            # 处理特殊字段
            if key == "colors" and cell_value:
                # 逗号分隔的颜色列表
                item[key] = [c.strip() for c in str(cell_value).split(",") if c.strip()]
            elif key == "sizes" and cell_value:
                # 逗号分隔的尺码列表
                item[key] = [s.strip() for s in str(cell_value).split(",") if s.strip()]
            elif key == "sku_images" and cell_value:
                # SKU图片：可以是JSON字符串或文件夹路径
                try:
                    import json
                    item[key] = json.loads(str(cell_value))
                except:
                    # 如果不是JSON，当作文件夹路径处理
                    item[key] = str(cell_value)
            else:
                item[key] = cell_value
        
        # 至少有标题或图片路径才算有效数据
        if item.get("image_path") or item.get("title"):
            data.append(item)

    wb.close()
    logger.info(f"从 {filepath} 读取了 {len(data)} 条上架数据")
    return data


def write_product_data(
    filepath: str,
    headers: list[str],
    product: dict,
    params: Optional[dict] = None,
    image_paths: Optional[dict] = None,
):
    """将采集到的商品数据追加写入 Excel（按照影刀上架参数格式）
    
    格式说明：
    - A-O列：按照影刀上架参数格式
    - P列：留空
    - S列开始：其他参数（商品链接、抓取时间等）
    """
    try:
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            # 写入表头（A-O列 + P列空 + S列开始的其他列）
            header_row = [
                "A", "B", "C平台加补后价格", "D", "E", 
                "F风格", "G颜色分类", "H适应季节", "I材质成分", "J面料",
                "K安全等级", "L", "M", "N主图路径", "O商品名称",
                "",  # P列留空
                "Q", "R",
                "S商品链接", "T抓取时间", "USKU图片路径"
            ]
            ws.append(header_row)

        params = params or {}
        image_paths = image_paths or {}

        # 按照列顺序填充数据
        row_data = [
            "",  # A列
            "",  # B列
            product.get("价格", ""),  # C列：平台加补后价格
            "",  # D列
            "",  # E列
            params.get("风格", ""),  # F列：风格
            params.get("颜色分类", "").replace(",", " ").replace("，", " "),  # G列：颜色分类
            params.get("适用季节", ""),  # H列：适应季节
            params.get("材质成分", ""),  # I列：材质成分
            params.get("面料", ""),  # J列：面料
            params.get("安全等级", ""),  # K列：安全等级
            "",  # L列
            "",  # M列
            image_paths.get("main", ""),  # N列：主图路径
            product.get("商品名称", ""),  # O列：商品名称
            "",  # P列：留空
            "",  # Q列
            "",  # R列
            product.get("商品链接", ""),  # S列：商品链接
            product.get("抓取时间", ""),  # T列：抓取时间
            image_paths.get("sku", ""),  # U列：SKU图片路径
        ]
        ws.append(row_data)
        wb.save(filepath)
        logger.info(f"数据已保存到: {filepath}")
    except Exception as e:
        logger.error(f"写入 Excel 失败: {e}")
